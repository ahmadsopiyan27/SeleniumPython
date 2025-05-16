import openpyxl



def get_data():

    final_list = []
    workbook = openpyxl.load_workbook("c:\AS\FD")
    sheet = workbook['LoginTest']
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range (2, total_rows+1):
        row_list =[]
        for c in range (1,total_cols+1):
            row_list.append(sheet.cell(row=r,column=c).value)
            final_list.append(row_list)

    return final_list


@pytest.mark.parametrize("username,password", get_data())
def test_login(username,password):
    print("Loggedn in using username: "+username+"and password"+password)